"""
Excel File Handler
Translates cell values in .xlsx/.xls files while preserving:
- Cell formatting (fonts, colors, borders, alignment)
- Merged cells
- Formulas (not translated, preserved as-is)
- Number formats
- Sheet names
- Multiple sheets
"""

from openpyxl import load_workbook
from core.utils import is_translatable


class ExcelHandler:
    """Handles translation of .xlsx and .xls files."""

    def translate(self, input_path, output_path, translator, progress_callback=None):
        """Translate all text content in an Excel file."""
        wb = load_workbook(input_path)

        # Count total cells for progress
        total_cells = sum(
            ws.max_row * ws.max_column
            for ws in wb.worksheets
            if ws.max_row and ws.max_column
        )
        total_cells = max(total_cells, 1)
        processed = 0

        for ws in wb.worksheets:
            # Translate sheet name
            if is_translatable(ws.title):
                ws.title = translator.translate_text(ws.title)

            # Get merged cell ranges (we only modify top-left cell)
            merged_coords = set()
            for merged_range in ws.merged_cells.ranges:
                min_col = merged_range.min_col
                min_row = merged_range.min_row
                merged_coords.add((min_row, min_col))
                # Other cells in the merged range should be skipped for writing
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        if row != min_row or col != min_col:
                            merged_coords.add((row, col))

            # Translate cells
            for row in ws.iter_rows():
                for cell in row:
                    coord = (cell.row, cell.column)

                    # Skip cells that are part of merged range but not top-left
                    if coord in merged_coords and cell.value is None:
                        processed += 1
                        continue

                    # Only translate string values
                    if cell.value and isinstance(cell.value, str) and is_translatable(cell.value):
                        cell.value = translator.translate_text(cell.value)

                    processed += 1

                    if progress_callback and processed % 100 == 0:
                        progress_callback(
                            processed / total_cells,
                            f"Translating sheet '{ws.title}'... ({processed}/{total_cells} cells)"
                        )

        wb.save(output_path)

        if progress_callback:
            progress_callback(1.0, "Excel translation complete!")